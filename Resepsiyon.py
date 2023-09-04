# sübhan akbenli    
import sqlite3
from PyQt5.QtWidgets import *
import sys
from datetime import datetime,timedelta

from anaekran_ui import * 
Uygulama = QApplication(sys.argv)
anaekran_main_window = QMainWindow()
anaekran_ui = Ui_MainWindow()
anaekran_ui.setupUi(anaekran_main_window)
anaekran_main_window.show()

from rezervasyonUI import *
rezervasyon_main_window = QMainWindow()
rezervasyon_ui = Ui_odaarama_MainWindow()
rezervasyon_ui.setupUi(rezervasyon_main_window)

from odalarUI import *
odalar_main_window=QMainWindow()
odalar_ui = Ui_odalar_MainWindow()
odalar_ui.setupUi(odalar_main_window)

from odaBilgi_ui import *
odaBilgi_main_window=QMainWindow()
odaBilgi_ui = Ui_odaBilgi_MainWindow()
odaBilgi_ui.setupUi(odaBilgi_main_window)

from musteriEkle_ui import *
musteri_main_window=QMainWindow()
musteri_ui=Ui_musteriekle_MainWindow()
musteri_ui.setupUi(musteri_main_window)

from odalarfiyat_ui import *
odalarfiyat_main_window=QMainWindow()
odalarfiyat_ui=Ui_odalarfiyat_MainWindow()
odalarfiyat_ui.setupUi(odalarfiyat_main_window)

from paketler_ui import *
paketler_main_window=QMainWindow()
paketler_ui=Ui_paketler_MainWindow()
paketler_ui.setupUi(paketler_main_window)

from musteriBilgiGuncelle_ui import * 
guncelle_main_window=QMainWindow()
guncelle_ui=Ui_musteriguncelle_MainWindow()
guncelle_ui.setupUi(guncelle_main_window)

from musteriHesap_ui import *
hesap_main_window=QMainWindow()
hesap_ui=Ui_hesap_MainWindow()
hesap_ui.setupUi(hesap_main_window)


from Rapor_ui import *
rapor_main_window=QMainWindow()
rapor_ui=Ui_rapor_MainWindow()
rapor_ui.setupUi(rapor_main_window)

con=sqlite3.connect("database.db")
curs=con.cursor()

curs.execute("CREATE TABLE IF NOT EXISTS odalar(haus Text,blok TEXT, odaAdi Text,kat Text,kapasite INT,fiyat INT)") #her bir bloktaki odalar ve doluluk oranı  
curs.execute("CREATE TABLE IF NOT EXISTS kisiler(blok Text, odaAdi Text,kat Text,kapasite Text,musteriNo INTEGER PRIMARY KEY AUTOINCREMENT,\
    ad Text,soyad Text,kategori Text,giris Date,cikis Date,telefon Text,mail Text,postaKodu Text,sehir Text,adres Text,tutar Int,odendi Text)") #her bir odadaki kişiler ve bilgileri  14
curs.execute("CREATE TABLE IF NOT EXISTS hesaplar(blok Text, odaAdi Text,musteriNo INT,HarcamaNo INTEGER PRIMARY KEY AUTOINCREMENT,ad Text,soyad Text,kategori Text,harcamatarihi Date,paketAdi Text,tutar Int,odendi Text)") #kişilerin ödemesi gereken tutar ve harcamalar
curs.execute("CREATE TABLE IF NOT EXISTS paketler(paketAdi Text ,fiyat Int)")
curs.execute("CREATE TABLE IF NOT EXISTS kategoriler(kategori Text)")



def musteri_ara(islem):
    if islem==0:
        giris,cikis=giris_cikistarihleri(rezervasyon_ui)
        if giris!= False:
    
            curs.execute("SELECT blok,odaAdi,musteriNo,ad,soyad,giris,cikis,telefon,tutar,odendi FROM kisiler Where giris <= ? and cikis > ?",(cikis,giris))
            data=curs.fetchall()
            odaArama_musteri_listele(data)
    elif islem==1:
        rzvNo=rezervasyon_ui.rzvNo_lineEdit.text()
        curs.execute("SELECT blok,odaAdi,musteriNo,ad,soyad,giris,cikis,telefon,tutar,odendi From kisiler WHERE (musteriNo LIKE {}) or (ad LIKE {}) or (soyad LIKE {}) ".format("'%"+rzvNo+"%'","'%"+rzvNo+"%'","'%"+rzvNo+"%'"))
        data=curs.fetchall()
        odaArama_musteri_listele(data)
    elif islem==2:
        curs.execute("SELECT blok,odaAdi,musteriNo,ad,soyad,giris,cikis,telefon,tutar,odendi From kisiler ")

        data=curs.fetchall()
        odaArama_musteri_listele(data)

def odaArama_musteri_listele(data):
    rezervasyon_ui.musteri_tableWidget.clearContents()
    row=0
    rezervasyon_ui.musteri_tableWidget.setRowCount(len(data))

    for blok,odaAdi,musteriNo,ad,soyad,giris,cikis,telefon,tutar,odendi in data:
        rezervasyon_ui.musteri_tableWidget.setItem(row,0,QTableWidgetItem(str(blok)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,1,QTableWidgetItem(str(odaAdi)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,2,QTableWidgetItem(str(musteriNo)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,3,QTableWidgetItem(str(ad+" "+soyad)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,4,QTableWidgetItem(str(giris)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,5,QTableWidgetItem(str(cikis)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,6,QTableWidgetItem(str(telefon)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,7,QTableWidgetItem(str(tutar)))
        rezervasyon_ui.musteri_tableWidget.setItem(row,8,QTableWidgetItem(str(odendi)))
        if odendi=="Ödendi":
            tablo_row_renklendir(rezervasyon_ui.musteri_tableWidget,row,QtGui.QColor(185, 250, 250))
        else:
            tablo_row_renklendir(rezervasyon_ui.musteri_tableWidget,row,QtGui.QColor(255,160,160))

        row+=1
    rezervasyon_ui.tabWidget.setCurrentIndex(1)

def oda_ara(ui):
    giris,cikis=giris_cikistarihleri(rezervasyon_ui)
    if giris!= False:
        curs.execute("SELECT blok,odaAdi,kat,kapasite FROM kisiler Where giris <= ? and cikis > ?",(cikis,giris))
        rezervasyon_ui.Tarihbaslik_label.setText(f"{giris} / {cikis}")

        data=curs.fetchall()
        curs.execute("SELECT blok,odaAdi,kat,kapasite FROM odalar")
        gonderilecek_data=curs.fetchall()

        sozluk={}
        for blok,odaAdi,kat,kapasite in data: #↓ burada odaların doluluk oranını öğreniyoruz
            try:
                kisi=sozluk[(blok,odaAdi)]
                sozluk[(blok,odaAdi)]=kisi+1
            except:
                sozluk[(blok,odaAdi)]=1
        rezervasyon_ui.tabWidget.setCurrentIndex(0)
        ui.odalar_tableWidget.clearContents()    

        odalistele(ui.odalar_tableWidget,gonderilecek_data,sozluk,
        ui.dolu_checkBox.isChecked(),ui.bos_checkBox.isChecked(),ui.yaridolu_checkBox.isChecked(),
        rezervasyon_ui.kat_eg_odalar_checkBox.isChecked(),rezervasyon_ui.kat_kg_odalar_checkBox.isChecked(),rezervasyon_ui.kat_1og_odalar_checkBox.isChecked(),rezervasyon_ui.kat_2og_odalar_checkBox.isChecked())

def hızlı_rezervasyonYap(ui):
    if ui ==odalar_ui:
        liste=odalar_ui.odalarTablosu.selectedItems()
    else:
        liste=rezervasyon_ui.odalar_tableWidget.selectedItems() 
    if len(liste)==0:
        ui.bilgilendirme.setText("Öncelikle bir tablodan oda seçiniz")
    else:
        tarihbaslik=ui.Tarihbaslik_label.text()
        blok=liste[0].text()
        odaAdi=liste[1].text()
        blokveodaAdi=f"{blok} / {odaAdi}"
        ui.bilgilendirme.setText("--")
        musteriekle_ac(tarihbaslik,blokveodaAdi)
        

def odalar_blok_getir(blok,kat):
    odalar_main_window.setWindowTitle(str(blok))
    odalar_main_window.show()
    giris,cikis=giris_cikistarihleri(odalar_ui)
    if giris!=False:
        
        odalar_ui.Tarihbaslik_label.setText(f"{giris} / {cikis}")

        if kat!=-1:
            curs.execute("SELECT blok,odaAdi,kat,kapasite FROM kisiler Where blok =? and kat = ? and giris <= ? and cikis > ?",(blok,kat,cikis,giris))
            data=curs.fetchall()
            curs.execute("SELECT blok,odaAdi,kat,kapasite FROM odalar Where blok =?",(blok,))
            gonderilecek_data=curs.fetchall()
        
        if kat ==-1:
            curs.execute("SELECT blok,odaAdi,kat,kapasite FROM kisiler Where blok =? and giris <= ? and cikis > ?",(blok,cikis,giris))
            data=curs.fetchall()
            curs.execute("SELECT blok,odaAdi,kat,kapasite FROM odalar Where blok =?",(blok,))
            gonderilecek_data=curs.fetchall()

        sozluk={}
        for blok,odaAdi,kat,kapasite in data: #↓ burada odaların doluluk oranını öğreniyoruz
            try:
                kisi=sozluk[(blok,odaAdi)]
                sozluk[(blok,odaAdi)]=kisi+1
            except:
                sozluk[(blok,odaAdi)]=1
        odalar_ui.odalarTablosu.clearContents()

        odalistele(odalar_ui.odalarTablosu,gonderilecek_data,sozluk,
                odalar_ui.doluodalar_checkBox.isChecked(),odalar_ui.tamamenbosodalar_checkBox.isChecked(),odalar_ui.yaridolu_odalar_checkBox.isChecked(),
                odalar_ui.kat_eg_odalar_checkBox.isChecked(),odalar_ui.kat_kg_odalar_checkBox.isChecked(),odalar_ui.kat_1og_odalar_checkBox.isChecked(),odalar_ui.kat_2og_odalar_checkBox.isChecked())


def odalistele(tablo,data,sozluk,
               doluGoster,bosGoster,yaridoluGoster,
               kat_eg=True,kat_kg=True,kat_1og=True,kat_2og=True):
    kontrol_liste=[]
    katlar=[]
    if kat_eg:
        katlar.append("EG")
    if kat_kg:
        katlar.append("KG")
    if kat_1og:
        katlar.append("1. OG")
    if kat_2og:
        katlar.append("2. OG")
    row=0

    for blok,odaAdi,kat,kapasite in data:
        tablo.setRowCount(row+1)
        if kat not in katlar:
            continue
        if ((blok,odaAdi,kat,kapasite) in kontrol_liste):
            continue
        kontrol_liste.append((blok,odaAdi,kat,kapasite))
        try:
            doluyatak=sozluk[((blok,odaAdi))]
        except:     doluyatak=0
        
        bosyatak=int(kapasite)-doluyatak

        if bosyatak<=0 and doluGoster==True:                        tablo.setItem(row,6,QTableWidgetItem(str("ODA TAMAMEN DOLU")))

        elif bosyatak==int(kapasite) and bosGoster==True:           tablo.setItem(row,6,QTableWidgetItem(str("ODA TAMAMEN BOŞ")))
    
        elif 0<bosyatak<int(kapasite) and yaridoluGoster==True:     tablo.setItem(row,6,QTableWidgetItem(str("BOŞ YERLER VAR")))
        
        else: continue

        tablo.setItem(row,0,QTableWidgetItem(str(blok)))
        tablo.setItem(row,1,QTableWidgetItem(str(odaAdi)))
        tablo.setItem(row,2,QTableWidgetItem(str(kat)))
        tablo.setItem(row,3,QTableWidgetItem(str(kapasite)))
        tablo.setItem(row,4,QTableWidgetItem(str(doluyatak)))
        tablo.setItem(row,5,QTableWidgetItem(str(bosyatak)))

        if bosyatak<=0 and doluGoster==True:                               tablo_row_renklendir(tablo,row,QtGui.QColor(225,45,45))

        elif bosyatak==int(kapasite) and bosGoster==True:                  tablo_row_renklendir(tablo,row,QtGui.QColor(45,225,45))
    
        elif 0<bosyatak<int(kapasite) and yaridoluGoster==True:            tablo_row_renklendir(tablo,row,QtGui.QColor(255,170,120))
        row+=1
    tablo.setRowCount(row)



def odabilgi_musteri_listele(data):
    odaBilgi_ui.odalarTablosu.clearContents()
    row=0
    odaBilgi_ui.odalarTablosu.setRowCount(len(data))
    for musteriNo,ad,soyad,giris,cikis,telefon,tutar,odendi in data:
        odaBilgi_ui.odalarTablosu.setItem(row,0,QTableWidgetItem(str(musteriNo)))
        odaBilgi_ui.odalarTablosu.setItem(row,1,QTableWidgetItem(str(ad+" "+soyad)))
        odaBilgi_ui.odalarTablosu.setItem(row,2,QTableWidgetItem(str(giris)))
        odaBilgi_ui.odalarTablosu.setItem(row,3,QTableWidgetItem(str(cikis)))
        odaBilgi_ui.odalarTablosu.setItem(row,4,QTableWidgetItem(str(telefon)))
        odaBilgi_ui.odalarTablosu.setItem(row,5,QTableWidgetItem(str(tutar)))
        odaBilgi_ui.odalarTablosu.setItem(row,6,QTableWidgetItem(str(odendi)))
        if odendi=="Ödendi":
            tablo_row_renklendir(odaBilgi_ui.odalarTablosu,row,QtGui.QColor(185,250,250))
        else:
            tablo_row_renklendir(odaBilgi_ui.odalarTablosu,row,QtGui.QColor(255,160,160))
        row+=1

def odabilgi_odagetir(tarihbaslik,liste):
    girisvecikis=tarihbaslik.split(" / ")
    giris=girisvecikis[0]
    cikis=girisvecikis[1]

    odaAdi=liste[1].text()
    odaBilgi_main_window.setWindowTitle(str(odaAdi))
    odaBilgi_main_window.show()
    blok=liste[0].text()

    kat=liste[2].text()
    kapasite=liste[3].text()
    dolu_yatak=liste[4].text()
    bos_yatak=liste[5].text()
    odaBilgi_ui.Tarihbaslik_label.setText(f"{tarihbaslik}")
    odaBilgi_ui.baslik_label.setText(f"{blok} / {odaAdi}")
    
    curs.execute("SELECT musteriNo,ad,soyad,giris,cikis,telefon,tutar,odendi FROM kisiler Where blok = ? and odaAdi= ? and giris <= ? and cikis > ?",(blok,odaAdi,cikis,giris))
    data=curs.fetchall()
    odabilgi_musteri_listele(data)

def musteri_musteriekle():
    ad=musteri_ui.ad_lineEdit.text()
    if len(ad)!=0:
        kategori=musteri_ui.kategori_comboBox.currentText()
        curs.execute("SELECT * from kategoriler where kategori = ?",(kategori,))
        data=curs.fetchall()
        if len(data)==0:
            curs.execute("INSERT INTO kategoriler values(?)",(kategori,))
            con.commit()
        soyad=musteri_ui.soyad_lineEdit.text()
        telefon=musteri_ui.telefon_lineEdit.text()
        mail=musteri_ui.mail_lineEdit.text()
        postaKodu=musteri_ui.postaKodu_lineEdit.text()
        sehir=musteri_ui.sehir_lineEdit.text()
        adres=musteri_ui.adres_textEdit.toPlainText()

        girisvecikis=musteri_ui.Tarihbaslik_label.text().split("/")
        giris=girisvecikis[0].strip()
        cikis=girisvecikis[1].strip()
        girislistesi=giris.split("-")
        cikislistesi=cikis.split("-")
        a=datetime(int(cikislistesi[0]),int(cikislistesi[1]),int(cikislistesi[2]))
        b=datetime(int(girislistesi[0]),int(girislistesi[1]),int(girislistesi[2]))
        gun_farki=(a-b).days
        blokveoda=musteri_ui.baslik_label.text().split(" / ")
        blok=blokveoda[0]
        odaAdi=blokveoda[1]
        curs.execute("SELECT * from odalar where blok = ? and odaAdi = ?",(blok,odaAdi))
        data=curs.fetchall()[0]
        fiyat=int(data[-1])*gun_farki
        curs.execute("INSERT INTO kisiler (blok,odaAdi,kat,kapasite,ad,soyad,kategori,giris,cikis,telefon,mail,postaKodu,sehir,adres,odendi) Values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (blok,odaAdi,data[-3],data[-2],ad,soyad,kategori,giris,cikis,telefon,mail,postaKodu,sehir,adres,"Ödenmedi"))
        con.commit()
        curs.execute("SELECT * FROM kisiler ORDER BY musteriNo DESC LIMIT 1")

        musteriNo = curs.fetchone()[4]
        curs.execute("INSERT INTO hesaplar (blok,odaAdi,musteriNo,ad,soyad,kategori,harcamatarihi,paketAdi,tutar,odendi) values(?,?,?,?,?,?,?,?,?,?)",(blok,odaAdi,musteriNo,ad,soyad,kategori,giris,f"{odaAdi}*{gun_farki}",fiyat,"ödenmedi"))
        con.commit()

        curs.execute("UPDATE kisiler SET tutar =? WHERE musteriNo=?",
            (fiyat,musteriNo))
        musteri_ui.bilgilendirme.setText(f"+Başarıyla Kaydedildi+ {ad}-{soyad}")
        musteri_ui.ad_lineEdit.clear()
        musteri_ui.soyad_lineEdit.clear()
        musteri_ui.telefon_lineEdit.clear()
        musteri_ui.mail_lineEdit.clear()
        musteri_ui.postaKodu_lineEdit.clear()
        musteri_ui.sehir_lineEdit.clear()
        musteri_ui.adres_textEdit.clear()
    else:
        musteri_ui.bilgilendirme.setText("müşteri adı girilmesi zorunludur")

def tablo_row_renklendir(tablo,row,color):
    for col in range(tablo.columnCount()+1):
        item = tablo.item(row, col)
        if item:
            item.setBackground(color)



def odalarfiyat_getir(blok=None):
    odalarfiyat_main_window.show()
    if blok==None:
        curs.execute("SELECT * from odalar",)
        data=curs.fetchall()
    odalarfiyat_listele(data)

def odalarfiyat_listele(data):
    odalarfiyat_ui.odalarTablosu.clearContents()
    row=0
    odalarfiyat_ui.odalarTablosu.setRowCount(len(data))
    for haus,blok,odaAdi,kat,kapasite,fiyat in data:  
        odalarfiyat_ui.odalarTablosu.setItem(row,0,QTableWidgetItem(str(blok)))
        odalarfiyat_ui.odalarTablosu.setItem(row,1,QTableWidgetItem(str(odaAdi)))
        odalarfiyat_ui.odalarTablosu.setItem(row,2,QTableWidgetItem(str(kat)))
        odalarfiyat_ui.odalarTablosu.setItem(row,3,QTableWidgetItem(str(kapasite)))
        odalarfiyat_ui.odalarTablosu.setItem(row,4,QTableWidgetItem(str(fiyat)))
        tablo_row_renklendir(odalarfiyat_ui.odalarTablosu,row,QtGui.QColor(185, 250, 250))
        row+=1

def odalarfiyat_paketgüncelle():
    odafiyat=odalarfiyat_ui.fiyat_lineEdit.text()
    try:
        odafiyat=int(odafiyat)
        rowlist=[]
        liste=odalarfiyat_ui.odalarTablosu.selectedIndexes()
        for row in liste:
            row=row.row()
            if row in rowlist:continue
            else: rowlist.append(row)
            blok=odalarfiyat_ui.odalarTablosu.item(row,0).text()
            odaAdi=odalarfiyat_ui.odalarTablosu.item(row,1).text()
            curs.execute("UPDATE odalar SET fiyat = ? WHERE blok = ? and odaAdi =?",(odafiyat,blok,odaAdi))
            con.commit()
            odalarfiyat_ui.bilgilendirme.setText("Oda günlük fiyatları başarıyla güncellendi")
        odalarfiyat_getir()

    except:
        odalarfiyat_ui.bilgilendirme.setText("!!! fiyatı kontrol ediniz !!!")



def paketler_getir():
    paketler_main_window.show()
    curs.execute("SELECT * From paketler")
    data=curs.fetchall()
    paketler_listele(data)

def paketler_listele(data):
    paketler_ui.odalarTablosu.clearContents()
    row=0
    paketler_ui.odalarTablosu.setRowCount(len(data))
    for paketAdi,fiyat in data:  
        paketler_ui.odalarTablosu.setItem(row,0,QTableWidgetItem(str(paketAdi)))
        paketler_ui.odalarTablosu.setItem(row,1,QTableWidgetItem(str(fiyat)))
        tablo_row_renklendir(paketler_ui.odalarTablosu,row,QtGui.QColor(245, 250, 150))
        
        row+=1

def paketler_paketEkle(paketAdi,fiyat): 
    if len(paketAdi)==0 or len(fiyat)==0: paketler_ui.bilgilendirme.setText("Lütfen paket adını ve fiyatını kontrol edin")
    else:
        try: 
            fiyat=int(fiyat)
            curs.execute("SELECT * FROM paketler where paketAdi = ?",(paketAdi,))
            data=curs.fetchall()
            if len(data)==0:
                curs.execute("INSERT INTO paketler values (?,?)",(paketAdi,fiyat))
                con.commit()
                paketler_ui.bilgilendirme.setText("Paket başarıyla eklendi")
                paketler_getir()
            else:paketler_ui.bilgilendirme.setText("Bu paket adı daha önceden eklendi")
        
        
        except: paketler_ui.bilgilendirme.setText("Lütfen paket fiyatını kontrol edin")

def paketler_paketgüncelle():
    paketfiyat=paketler_ui.fiyat_lineEdit.text()
    try:
        paketfiyat=int(paketfiyat)
        rowlist=[]
        liste=paketler_ui.odalarTablosu.selectedIndexes()
        for row in liste:
            row=row.row()
            if row in rowlist:continue
            else: rowlist.append(row)
            paketAdi=paketler_ui.odalarTablosu.item(row,0).text()
            curs.execute("UPDATE paketler SET fiyat = ? WHERE paketAdi = ?",(paketfiyat,paketAdi))
            con.commit()
            paketler_ui.bilgilendirme.setText("Paketler başarıyla güncellendi")
        paketler_getir()

    except:
        paketler_ui.bilgilendirme.setText("!!! Paket fiyatını kontrol ediniz !!!")

def paketler_sil():
    try:
        rowlist=[]
        liste=paketler_ui.odalarTablosu.selectedIndexes()
        for row in liste:
            row=row.row()
            if row in rowlist:continue
            else: rowlist.append(row)
            paketAdi=paketler_ui.odalarTablosu.item(row,0).text()
            curs.execute("DELETE From paketler WHERE paketAdi = ?",(paketAdi,))
            con.commit()
            paketler_ui.bilgilendirme.setText("Paketler başarıyla silindi")
        paketler_getir()

    except:
        paketler_ui.bilgilendirme.setText("!!! Bir hata oluştu !!!")



def guncelle_musterigetir(ui):
    musteriNo=None

    if ui==rezervasyon_ui:
        try:        musteriNo=rezervasyon_ui.musteri_tableWidget.selectedItems()[2].text()
        except:     rezervasyon_ui.bilgilendirme.setText("Lütfen öncelikle müşteriler tablosundan bir müştreri seçiniz")
    if ui==odaBilgi_ui:
        try:            musteriNo=odaBilgi_ui.odalarTablosu.selectedItems()[0].text()
        except:         odaBilgi_ui.bilgilendirme.setText("Lütfen öncelikle müşteriler tablosundan bir müştreri seçiniz")

    if musteriNo!=None:
        guncelle_main_window.show()
        curs.execute("SELECT blok,odaAdi,ad,soyad,giris,cikis,telefon,mail,postaKodu,sehir,adres FROM  kisiler Where musteriNo = ?",(musteriNo,))
        data=curs.fetchall()
        guncelle_main_window.setWindowTitle(f"Güncelle -{musteriNo}")
        for blok,odaAdi,ad,soyad,giris,cikis,telefon,mail,postaKodu,sehir,adres in data:
            guncelle_ui.Tarihbaslik_label.setText(f"{giris} / {cikis}")
            guncelle_ui.baslik_label.setText(f"{blok} / {odaAdi}")
            guncelle_ui.ad_lineEdit.setText(str(ad))
            guncelle_ui.soyad_lineEdit.setText(str(soyad))
            guncelle_ui.telefon_lineEdit.setText(str(telefon))
            guncelle_ui.mail_lineEdit.setText(str(mail))
            guncelle_ui.postaKodu_lineEdit.setText(str(postaKodu))
            guncelle_ui.sehir_lineEdit.setText(str(sehir))
            guncelle_ui.adres_textEdit.setText(str(adres))
        guncelle_ui.bilgilendirme.setText("Müşteri Bilgisi başarıyla güncellendi")

def guncelle_musteriguncelle():
    musteriNo=int(guncelle_main_window.windowTitle().split("-")[1])
    ad=guncelle_ui.ad_lineEdit.text()
    soyad=guncelle_ui.soyad_lineEdit.text()
    telefon=guncelle_ui.telefon_lineEdit.text()
    mail=guncelle_ui.mail_lineEdit.text()
    postakodu=guncelle_ui.postaKodu_lineEdit.text()
    sehir=guncelle_ui.sehir_lineEdit.text()
    adres=guncelle_ui.adres_textEdit.toPlainText()
    curs.execute("UPDATE kisiler SET ad=?, soyad=?, telefon=?, mail=?, postaKodu=?, sehir=?, adres=? WHERE musteriNo=?",
                (ad,soyad,telefon,mail,postakodu,sehir,adres,musteriNo))
    con.commit()
    guncelle_ui.bilgilendirme.setText("Müşteri kaydı başarıyla güncellendi")




def odaarama_ac():
    rezervasyon_main_window.show()

def musteriekle_ac(tarih,blokveoda):
    musteri_main_window.show()
    musteri_main_window.setWindowTitle(str(blokveoda))
    musteri_ui.baslik_label.setText(blokveoda)
    musteri_ui.Tarihbaslik_label.setText(tarih)
    musteri_ui.kategori_comboBox.clear()
    curs.execute("SELECT * from kategoriler")
    data=curs.fetchall()
    for kategori in data:
        kategori=kategori[0]
        musteri_ui.kategori_comboBox.addItem(kategori)

def musteriekle_kategoriEkle():
    text=musteri_ui.kategori_lineEdit.text()
    if len(text)>2:
        musteri_ui.kategori_comboBox.addItem(text)
        musteri_ui.kategori_comboBox.setCurrentText(text)
        musteri_ui.bilgilendirme.setText("")

    else:
        musteri_ui.bilgilendirme.setText("Kategori Adı çok kısa")

def musteri_sil(ui,liste):
    if len(liste)==0:
        ui.bilgilendirme.setText("Tablodan bir satır seçiniz")
    else:
        if ui==rezervasyon_ui:
            musteriNo=liste[2].text()
        elif ui==odaBilgi_ui:
            musteriNo=liste[0].text()

        curs.execute("DELETE FROM kisiler Where musteriNo = ? ",(musteriNo,))
        con.commit()
        ui.bilgilendirme.setText("Müşteri kaydı başarıyla silindi - "+musteriNo)
        if ui==rezervasyon_ui:
            musteri_ara(2)

def tarihleridüzenle(giris,cikis):
    rezervasyon_ui.giris_dateEdit.setDate(giris)
    odalar_ui.giris_dateEdit.setDate(giris)
    hesap_ui.harcama_dateEdit.setDate(giris)
    rapor_ui.giris_dateEdit.setDate(giris)
    rapor_ui.cikis_dateEdit.setDate(cikis)
    rezervasyon_ui.cikis_dateEdit.setDate(cikis)
    odalar_ui.cikis_dateEdit.setDate(cikis)

def giris_cikistarihleri(ui):
    if ui==rezervasyon_ui:
        giris=rezervasyon_ui.giris_dateEdit.date().toPyDate()
        cikis=rezervasyon_ui.cikis_dateEdit.date().toPyDate()
    elif ui ==odalar_ui:
        giris=odalar_ui.giris_dateEdit.date().toPyDate()
        cikis=odalar_ui.cikis_dateEdit.date().toPyDate()
    if cikis<=giris:
        ui.bilgilendirme.setText("Çıkış tarihi giriş tarihinden küçük veya aynı olamaz")
        return False,False
    else:
        ui.bilgilendirme.setText("--")
        return giris,cikis


def hesap_getir(musteriNo):
    try:
        hesap_ui.musteriNo_lineEdit.setText(str(musteriNo))
        musteriNo==int(musteriNo)
        curs.execute("SELECT blok,odaAdi,musteriNo,harcamaNo,ad,soyad,harcamatarihi,paketAdi,tutar,odendi FROM hesaplar WHERE musteriNo = ?",(musteriNo,))
        data=curs.fetchall()
        if len(data)==0:
            hesap_ui.bilgilendirme.setText("Müşteriye ait harcama bulunamadı")        
            hesap_main_window.show()  
        else:
            blok=data[0][0]
            odaAdi=data[0][1]
            ad=data[0][4]
            soyad=data[0][5]
            hesap_ui.baslik_label.setText(f"{blok} / {odaAdi} / {ad} {soyad}")
            curs.execute("SELECT giris,cikis FROM kisiler where musteriNo=?",(musteriNo,))
            giris,cikis=curs.fetchall()[0]
            hesap_ui.Tarihbaslik_label.setText(f"{giris} / {cikis}")
            hesap_hesaplistele(data)
            curs.execute("SELECT * From paketler")
            data=curs.fetchall()
            hesap_ui.paketlerTablosu.clearContents()
            row=0
            hesap_ui.paketlerTablosu.setRowCount(len(data))
            for paketAdi,fiyat in data:  
                hesap_ui.paketlerTablosu.setItem(row,0,QTableWidgetItem(str(paketAdi)))
                hesap_ui.paketlerTablosu.setItem(row,1,QTableWidgetItem(str(fiyat)))
                tablo_row_renklendir(hesap_ui.paketlerTablosu,row,QtGui.QColor(245, 250, 150))
                row+=1

    except:
        hesap_ui.bilgilendirme.setText("!!!Müşteri No sadece sayı olabilir !!!")
        hesap_main_window.show()

def hesap_hesaplistele(data):
    hesap_ui.hesaplarTablosu.clearContents()
    row=0
    hesap_ui.hesaplarTablosu.setRowCount(len(data))
    toplam=0
    musteriNo=0
    for blok,odaAdi,musteriNo,harcamaNo,ad,soyad,harcamatarihi,paketAdi,tutar,odendi in data:
        musteriNo=musteriNo
        hesap_ui.hesaplarTablosu.setItem(row,0,QTableWidgetItem(str(musteriNo)))
        hesap_ui.hesaplarTablosu.setItem(row,1,QTableWidgetItem(str(harcamaNo)))
        hesap_ui.hesaplarTablosu.setItem(row,2,QTableWidgetItem(str(harcamatarihi)))
        hesap_ui.hesaplarTablosu.setItem(row,3,QTableWidgetItem(str(paketAdi)))
        hesap_ui.hesaplarTablosu.setItem(row,4,QTableWidgetItem(str(tutar)))
        hesap_ui.hesaplarTablosu.setItem(row,5,QTableWidgetItem(str(odendi)))
        if odendi=="Ödendi":
            tablo_row_renklendir(hesap_ui.hesaplarTablosu,row,QtGui.QColor(155,155,255))
        else:
            toplam+=tutar
            tablo_row_renklendir(hesap_ui.hesaplarTablosu,row,QtGui.QColor(255,255,155))

        row+=1
    if toplam==0:   curs.execute("UPDATE kisiler SET tutar =?,odendi=? WHERE musteriNo=?",(toplam,"Ödendi",musteriNo))
    else:           curs.execute("UPDATE kisiler SET tutar =?,odendi=? WHERE musteriNo=?",(toplam,"Ödenmedi",musteriNo))
    con.commit()
    hesap_main_window.show()

def hesap_harcamaekle(liste):
    if len(liste)>=2:
        try:
            paketAdi=liste[0].text()
            fiyat=float(liste[1].text())
        except:
            paketAdi=liste[0]
            fiyat=liste[1]
        try:
            musteriNo=int(hesap_ui.musteriNo_lineEdit.text())
            harcamatarihi=hesap_ui.harcama_dateEdit.date().toPyDate()
            curs.execute("Select blok,odaAdi,ad,soyad,kategori FROM kisiler Where musteriNo =?",(musteriNo,))
            data=curs.fetchall()
            if len(data)!=0:
                for blok,odaAdi,ad,soyad,kategori in data:
                    curs.execute("INSERT INTO hesaplar (blok,odaAdi,musteriNo,ad,soyad,kategori,harcamatarihi,paketAdi,tutar,odendi) values(?,?,?,?,?,?,?,?,?,?)",
                                (blok,odaAdi,musteriNo,ad,soyad,kategori,harcamatarihi,paketAdi,fiyat,"ödenmedi"))
                    con.commit()
                hesap_ui.bilgilendirme.setText("harcama başarıyla eklendi")
                hesap_getir(musteriNo)
        except:
            hesap_ui.bilgilendirme.setText("MüşteriNo alanında sadece numerik tuşlar kullanılabilir")
    else:
        hesap_ui.bilgilendirme.setText("!!! Paket seçiniz")

    pass

def hesap_ozelharcamaekle():
    paketAdi=hesap_ui.paketEkleadi_lineEdit.text()
    if len(paketAdi)!=0:
        try:
            fiyat=float(hesap_ui.paketEklefiyat_lineEdit.text())
            hesap_harcamaekle([paketAdi,fiyat])

        except:hesap_ui.bilgilendirme.setText("!!!Fiyat sayı cinsinden olmalıdır!!!")
    else:hesap_ui.bilgilendirme.setText("!!! Paket Adını yazınız!!!")

def hesap_odendi():
    liste=hesap_ui.hesaplarTablosu.selectedItems()
    if len(liste)!=0:
        musteriNo=int(liste[0].text())
        harcamaNo=int(liste[1].text())
        odendi=liste[-1].text()
        if odendi=="Ödendi":
            curs.execute("UPDATE hesaplar SET odendi =? WHERE harcamaNo=?",("Ödenmedi",harcamaNo))
            con.commit()
            curs.execute("UPDATE kisiler SET odendi =? WHERE musteriNo=?",("Ödenmedi",musteriNo))
            con.commit()

        else:
            curs.execute("UPDATE hesaplar SET odendi =? WHERE harcamaNo=?",("Ödendi",harcamaNo))
            con.commit()
        hesap_getir(musteriNo)

def raporekrani_Ac():
    rapor_main_window.show()
    curs.execute("SELECT * from kategoriler")
    data=curs.fetchall()
    rapor_ui.kategori_comboBox.clear()
    rapor_ui.kategori_comboBox.addItem("--")

    for kategori in data:
        kategori=kategori[0]
        rapor_ui.kategori_comboBox.addItem(kategori)

def rapor_filtre(giris,cikis,aranankategori="--",arananblok="--"):
    text=""
    if arananblok!="--":
        arananblok=arananblok.capitalize()
        text+=f"blok = '{arananblok}' AND"
    print(arananblok)
    if aranankategori=="--":     curs.execute(f"SELECT blok,odaAdi,musteriNo,harcamaNo,ad,soyad,kategori,harcamatarihi,paketAdi,tutar,odendi FROM hesaplar WHERE {text} harcamatarihi BETWEEN ? AND ?",(giris,cikis))
    else:                  curs.execute(f"SELECT blok,odaAdi,musteriNo,harcamaNo,ad,soyad,kategori,harcamatarihi,paketAdi,tutar,odendi FROM hesaplar WHERE {text} kategori = ? and (harcamatarihi BETWEEN ? AND ?)",(aranankategori,giris,cikis))
    dataHesaplar=curs.fetchall()
        
    if aranankategori=="--":        curs.execute(f"SELECT blok,odaAdi,kat,kapasite,musteriNo,ad,soyad,kategori,giris,cikis,odendi FROM kisiler Where {text} giris BETWEEN ? AND ?",(giris,cikis))
    else:               curs.execute(f"SELECT blok,odaAdi,kat,kapasite,musteriNo,ad,soyad,kategori,giris,cikis,odendi FROM kisiler Where {text} kategori = ? and (giris BETWEEN ? AND ?)",(aranankategori,giris,cikis))
    dataKisiler=curs.fetchall()
    dosyaAdi=arananblok+" "+aranankategori+" "+str(giris)+" "+str(cikis)
    return dataHesaplar,dataKisiler,dosyaAdi

def rapor_tabloyaGetir(dataHesaplardataKisilerdosyaAdi):
    dataHesaplar,dataKisiler,dosyaAdi=dataHesaplardataKisilerdosyaAdi
    tutarsozluk={}
    paketsozluk={}
    bloksozluk={}
    for blok,odaAdi,musteriNo,HarcamaNo,ad,soyad,kategori,harcamatarihi,paketAdi,tutar,odendi in dataHesaplar:
        if odendi=="Ödendi":    odenen2=tutar
        else: odenen2=0
        
        try:
            odenen,toplam=bloksozluk[blok]
            bloksozluk[blok]=odenen+odenen2,toplam+tutar
        except:
            bloksozluk[blok]=odenen2,tutar

        try:
            adet,adettutar,odenen,toplam=paketsozluk[paketAdi]
            paketsozluk[paketAdi]=adet+1,tutar,odenen+odenen2,toplam+tutar
        except:
            paketsozluk[paketAdi]=1,tutar,odenen2,tutar

        try:
            odenen,toplam=tutarsozluk[musteriNo]
            tutarsozluk[musteriNo]=odenen+odenen2,toplam+tutar
        except:
            tutarsozluk[musteriNo]=odenen2,tutar
    
    print(bloksozluk.items())
    rapor_ui.paketToplam_tableWidget.setRowCount(len(paketsozluk.items())+1)
    row=0
    odenentoplami=0
    toplamtoplami=0
    for paketAdi,(adet,adettutar,odenen,toplam) in paketsozluk.items():
        rapor_ui.paketToplam_tableWidget.setItem(row,0,QTableWidgetItem(str(paketAdi)))
        rapor_ui.paketToplam_tableWidget.setItem(row,1,QTableWidgetItem(str(adet)))
        rapor_ui.paketToplam_tableWidget.setItem(row,2,QTableWidgetItem(str(adettutar)))
        rapor_ui.paketToplam_tableWidget.setItem(row,3,QTableWidgetItem(str(odenen)))
        odenentoplami+=odenen
        rapor_ui.paketToplam_tableWidget.setItem(row,4,QTableWidgetItem(str(toplam)))
        toplamtoplami+=toplam
        row+=1
    rapor_ui.paketToplam_tableWidget.setItem(row,3,QTableWidgetItem(str(odenentoplami)))
    rapor_ui.paketToplam_tableWidget.setItem(row,4,QTableWidgetItem(str(toplamtoplami)))

    rapor_ui.blokToplam_tableWidget.setRowCount(len(bloksozluk.items())+1)
    row=0
    odenentoplami=0
    toplamtoplami=0
    for blok,(odenen,toplam) in bloksozluk.items():
        rapor_ui.blokToplam_tableWidget.setItem(row,0,QTableWidgetItem(str(blok)))
        rapor_ui.blokToplam_tableWidget.setItem(row,1,QTableWidgetItem(str(odenen)))
        odenentoplami+=odenen
        rapor_ui.blokToplam_tableWidget.setItem(row,2,QTableWidgetItem(str(toplam)))
        toplamtoplami+=toplam
        row+=1

    rapor_ui.blokToplam_tableWidget.setItem(row,1,QTableWidgetItem(str(odenentoplami)))
    rapor_ui.blokToplam_tableWidget.setItem(row,2,QTableWidgetItem(str(toplamtoplami)))

    row=0
    odenentoplami=0
    toplamtoplami=0
    rapor_ui.musteri_tableWidget.setRowCount(len(dataKisiler)+1)

    for blok,odaAdi,kat,kapasite,musteriNo,ad,soyad,kategori,giris,cikis,odendi in dataKisiler:
        print(blok)
        odenen,toplam=tutarsozluk[musteriNo]
        rapor_ui.musteri_tableWidget.setItem(row,0,QTableWidgetItem(str(blok)))
        rapor_ui.musteri_tableWidget.setItem(row,1,QTableWidgetItem(str(odaAdi)))
        rapor_ui.musteri_tableWidget.setItem(row,2,QTableWidgetItem(str(musteriNo)))
        rapor_ui.musteri_tableWidget.setItem(row,3,QTableWidgetItem(str(ad+" "+soyad)))
        rapor_ui.musteri_tableWidget.setItem(row,4,QTableWidgetItem(str(giris)))
        rapor_ui.musteri_tableWidget.setItem(row,5,QTableWidgetItem(str(cikis)))
        rapor_ui.musteri_tableWidget.setItem(row,6,QTableWidgetItem(str(odenen)))
        rapor_ui.musteri_tableWidget.setItem(row,7,QTableWidgetItem(str(toplam)))
        odenentoplami+=odenen
        toplamtoplami+=toplam
        row+=1   
        print("++++++++++++++++++++++s")
    rapor_ui.musteri_tableWidget.setItem(row,6,QTableWidgetItem(str(odenentoplami)))
    rapor_ui.musteri_tableWidget.setItem(row,7,QTableWidgetItem(str(toplamtoplami)))

def ExceleYaz(dataHesaplardataKisilerdosyaAdi):
    dataHesaplar,dataKisiler,dosyaAdi=dataHesaplardataKisilerdosyaAdi
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill,Font
    # Arka plan rengini ayarlamak için doldurma nesnesi oluştur
    fillSari = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Burada sarı renk seçildi
    fillYesil = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Burada yeşil renk seçildi
    fillKirmizi = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Burada kırmızı renk seçildi
    fillMavi = PatternFill(start_color='66CCFF', end_color='66CCFF', fill_type='solid')  # Burada mavi renk seçildi
    

    workbook = Workbook()
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 25  # A sütunu için genişlik 15
    sheet.column_dimensions['B'].width = 25  # B sütunu için genişlik 20
    sheet.column_dimensions['C'].width = 20  # c sütunu için genişlik 20
    sheet.column_dimensions['D'].width = 25  # D sütunu için genişlik 20
    sheet.column_dimensions['E'].width = 15  # E sütunu için genişlik 20
    sheet.column_dimensions['F'].width = 20  # E sütunu için genişlik 20

    row=1
    font = Font(size=14)  # 14 punto metin boyutu
    sheet[f'A{row}'] = 'Blok/ Oda Adı'
    sheet[f'A{row}'].fill = fillMavi
    sheet[f'A{row}'].font = font
    
    sheet[f'B{row}'] = 'Müşteri No/ Ad Soyad'
    sheet[f'B{row}'].fill = fillMavi
    sheet[f'B{row}'].font = font
    
    sheet[f'C{row}'] = "Kategori"
    sheet[f'C{row}'].fill = fillMavi
    sheet[f'C{row}'].font = font

    sheet[f'D{row}'] = 'Paket Adı'
    sheet[f'D{row}'].fill = fillMavi
    sheet[f'D{row}'].font = font

    sheet[f'E{row}'] = 'Tutar'
    sheet[f'E{row}'].fill = fillMavi
    sheet[f'E{row}'].font = font
    
    sheet[f'F{row}'] = 'Harcama Tarihi'
    sheet[f'F{row}'].fill = fillMavi
    sheet[f'F{row}'].font = font


    row+=1
    tutarsozluk={}
    toplamhucre=0
    for blok,odaAdi,musteriNo,HarcamaNo,ad,soyad,kategori,harcamatarihi,paketAdi,tutar,odendi in dataHesaplar:
        if odendi=="Ödendi":        fill=fillYesil
        else:                       fill = fillKirmizi
        sheet[f'A{row}'] = f'{blok}/{odaAdi}'
        sheet[f'A{row}'].fill = fillSari
        
        sheet[f'B{row}'] = f'{musteriNo}/{ad} {soyad}'
        sheet[f'B{row}'].fill = fillSari
        
        sheet[f'C{row}'] = f'{kategori}'
        sheet[f'C{row}'].fill = fillSari

        sheet[f'D{row}'] = f'{paketAdi}'
        sheet[f'D{row}'].fill = fill

        sheet[f'E{row}'] = f'{tutar}'
        sheet[f'E{row}'].fill = fill
        toplamhucre+=tutar
        sheet[f'F{row}'] = f'{harcamatarihi}'
        sheet[f'F{row}'].fill = fill
        if odendi=="Ödendi":    odenen2=tutar
        else: odenen2=0
        try:
            odenen,toplam=tutarsozluk[musteriNo]
            tutarsozluk[musteriNo]=odenen+odenen2,toplam+tutar
        except:
            tutarsozluk[musteriNo]=odenen2,tutar
        row+=1
    sheet[f'E{row}'] = f'{toplamhucre}'
    sheet[f'E{row}'].fill = fillMavi
    
    sheet2=workbook.create_sheet("Kisiler")
    
    sheet2.column_dimensions['A'].width = 25  # A sütunu için genişlik 15
    sheet2.column_dimensions['B'].width = 25  # B sütunu için genişlik 20
    sheet2.column_dimensions['C'].width = 20  # c sütunu için genişlik 20
    sheet2.column_dimensions['D'].width = 25  # D sütunu için genişlik 20
    sheet2.column_dimensions['E'].width = 15  # E sütunu için genişlik 20
    sheet2.column_dimensions['F'].width = 20  # E sütunu için genişlik 20

    row=1
    font = Font(size=14)  # 14 punto metin boyutu
    sheet2[f'A{row}'] = 'Blok/ Oda Adı'
    sheet2[f'A{row}'].fill = fillMavi
    sheet2[f'A{row}'].font = font
    
    sheet2[f'B{row}'] = 'Müşteri No/ Ad Soyad'
    sheet2[f'B{row}'].fill = fillMavi
    sheet2[f'B{row}'].font = font
    
    sheet2[f'C{row}'] = "Kategori"
    sheet2[f'C{row}'].fill = fillMavi
    sheet2[f'C{row}'].font = font

    sheet2[f'D{row}'] = 'Giris/Çıkış'
    sheet2[f'D{row}'].fill = fillMavi
    sheet2[f'D{row}'].font = font

    sheet2[f'E{row}'] = 'Toplam'
    sheet2[f'E{row}'].fill = fillMavi
    sheet2[f'E{row}'].font = font

    sheet2[f'F{row}'] = 'Ödenen'
    sheet2[f'F{row}'].fill = fillMavi
    sheet2[f'F{row}'].font = font
   
    row+=1
    toplamtoplami=0
    odenentoplami=0
    for blok,odaAdi,kat,kapasite,musteriNo,ad,soyad,kategori,giris,cikis,odendi in dataKisiler:
        if odendi=="Ödendi":        fill=fillYesil
        else:                       fill = fillKirmizi
        
        odenen,toplam=tutarsozluk[musteriNo]
        
        sheet2[f'A{row}'] = f'{blok}/{odaAdi}'
        sheet2[f'A{row}'].fill = fillSari
        
        sheet2[f'B{row}'] = f'{musteriNo}/{ad} {soyad}'
        sheet2[f'B{row}'].fill = fillSari
        
        sheet2[f'C{row}'] = f'{kategori}'
        sheet2[f'C{row}'].fill = fillSari

        sheet2[f'D{row}'] = f'{giris}/{cikis}'
        sheet2[f'D{row}'].fill = fill

        sheet2[f'E{row}'] = f'{toplam}'
        sheet2[f'E{row}'].fill = fill
        toplamtoplami+=toplam
        sheet2[f'F{row}'] = f'{odenen}'
        sheet2[f'F{row}'].fill = fill   
        odenentoplami+=odenen
        row+=1   

    sheet2[f'E{row}'] = f'{toplamtoplami}'
    sheet2[f'E{row}'].fill = fillMavi
    sheet2[f'F{row}'] = f'{odenentoplami}'
    sheet2[f'F{row}'].fill = fillMavi   
    # Dosyayı kaydet
    try:

        workbook.save(f'Rapor {dosyaAdi}.xlsx')
    except: 
        print("Excel Dosyasını kapatınız")

def DolulukOranı(giris,cikis):
    
    liste =[anaekran_ui.blaueshaus_pushButton,   anaekran_ui.violeteshaus_pushButton,    anaekran_ui.orangeneshaus_pushButton,        anaekran_ui.gruneshaus_pushButton,
        anaekran_ui.roteshaus_pushButton,            anaekran_ui.provenholz_pushButton,    anaekran_ui.taubental_pushButton,            anaekran_ui.appartment5_pushButton,
        anaekran_ui.appartment13_pushButton,         anaekran_ui.beneckerlinde_pushButton]
    


    for button in liste:
        blok=button.text().replace("İ","i").replace("Ö","ö").capitalize()
        

        
        curs.execute("SELECT blok,odaAdi,kat,kapasite FROM kisiler Where blok =? and giris <= ? and cikis > ?",(blok,cikis,giris))
        doluAdet=len(curs.fetchall())
        curs.execute("SELECT SUM(kapasite) FROM odalar WHERE blok = ?", (blok,))
        kapasite=int(curs.fetchall()[0][0])
        button.setText(blok+" - "+f"{doluAdet}/{kapasite}")
        if kapasite==doluAdet:      button.setStyleSheet("background-color:red")


bugun = datetime.now()
bugunun_tarihi=bugun.date()
yarinin_tarihi = (bugun + timedelta(days=1)).date()
tarihfarki=(bugunun_tarihi-yarinin_tarihi).days
tarihleridüzenle(bugunun_tarihi,yarinin_tarihi)
DolulukOranı(bugunun_tarihi,yarinin_tarihi)

anaekran_ui.rezervasyonArama_pushButton.clicked.connect(lambda : odaarama_ac())
anaekran_ui.blaueshaus_pushButton.clicked.connect(lambda : odalar_blok_getir("Blaueshaus",-1))
anaekran_ui.violeteshaus_pushButton.clicked.connect(lambda : odalar_blok_getir("Violeteshaus",-1))
anaekran_ui.orangeneshaus_pushButton.clicked.connect(lambda : odalar_blok_getir("Orangeshaus",-1))
anaekran_ui.gruneshaus_pushButton.clicked.connect(lambda : odalar_blok_getir("Grüneshaus",-1))
anaekran_ui.roteshaus_pushButton.clicked.connect(lambda : odalar_blok_getir("Roteshaus",-1))
anaekran_ui.provenholz_pushButton.clicked.connect(lambda : odalar_blok_getir("Prövenholz",-1))
anaekran_ui.taubental_pushButton.clicked.connect(lambda : odalar_blok_getir("Taubental",-1))
anaekran_ui.appartment5_pushButton.clicked.connect(lambda : odalar_blok_getir("Appartment 5",-1))
anaekran_ui.appartment13_pushButton.clicked.connect(lambda : odalar_blok_getir("Appartment 13",-1))
anaekran_ui.beneckerlinde_pushButton.clicked.connect(lambda : odalar_blok_getir("Beneckerlinde",-1))
anaekran_ui.odafiyatguncelle_pushButton.clicked.connect(lambda : odalarfiyat_getir())
anaekran_ui.paketler_pushButton.clicked.connect(lambda : paketler_getir())
anaekran_ui.rapor_pushButton.clicked.connect(lambda : raporekrani_Ac())

odalar_ui.odalistele_pushButton.clicked.connect(lambda : odalar_blok_getir("Blaueshaus",-1))
odalar_ui.odalarTablosu.doubleClicked.connect(lambda :odabilgi_odagetir(odalar_ui.Tarihbaslik_label.text(),odalar_ui.odalarTablosu.selectedItems()))
odalar_ui.rezYap_pushButton.clicked.connect(lambda : hızlı_rezervasyonYap(odalar_ui) )

rezervasyon_ui.tarihileMusteriAra_pushButton.clicked.connect(lambda : musteri_ara(0))
rezervasyon_ui.musterirezAra_pushButton.clicked.connect(lambda: musteri_ara(1))
rezervasyon_ui.musterilerlistele_pushButton.clicked.connect(lambda : musteri_ara(2))
rezervasyon_ui.musterirezSil_pushButton.clicked.connect(lambda : 
                                                        musteri_sil(rezervasyon_ui, rezervasyon_ui.musteri_tableWidget.selectedItems()))
rezervasyon_ui.musterirezguncelle_pushButton.clicked.connect(lambda : guncelle_musterigetir(rezervasyon_ui))
rezervasyon_ui.rezAra_pushButton.clicked.connect(lambda :oda_ara(rezervasyon_ui))
rezervasyon_ui.odalar_tableWidget.doubleClicked.connect(lambda : odabilgi_odagetir(rezervasyon_ui.Tarihbaslik_label.text(),
                                                                                   rezervasyon_ui.odalar_tableWidget.selectedItems()))
rezervasyon_ui.rezYap_pushButton.clicked.connect(lambda : hızlı_rezervasyonYap(rezervasyon_ui))
rezervasyon_ui.musteri_tableWidget.doubleClicked.connect(lambda :hesap_getir(rezervasyon_ui.musteri_tableWidget.selectedItems()[2].text()))

odaBilgi_ui.musteriekle_pushButton.clicked.connect(lambda :musteriekle_ac(odaBilgi_ui.Tarihbaslik_label.text(),odaBilgi_ui.baslik_label.text()))
odaBilgi_ui.musterisil_pushButton.clicked.connect(lambda : musteri_sil(odaBilgi_ui,odaBilgi_ui.odalarTablosu.selectedItems()))
odaBilgi_ui.musteriguncelle_pushButton_2.clicked.connect(lambda : guncelle_musterigetir(odaBilgi_ui))
odaBilgi_ui.odalarTablosu.doubleClicked.connect(lambda : hesap_getir(odaBilgi_ui.odalarTablosu.selectedItems()[0].text()))


musteri_ui.musteriekle_pushButton.clicked.connect(lambda : musteri_musteriekle())
musteri_ui.kategoriekle_pushButton.clicked.connect(lambda : musteriekle_kategoriEkle())
paketler_ui.paketEkle_pushButton.clicked.connect(lambda : paketler_paketEkle(paketler_ui.paketEkleadi_lineEdit.text(),paketler_ui.paketEklefiyat_lineEdit.text()))
paketler_ui.paketfiyatkaydet_pushButton.clicked.connect(lambda : paketler_paketgüncelle())
paketler_ui.paketSil_pushButton.clicked.connect(lambda : paketler_sil())

odalarfiyat_ui.fiyatkaydet_pushButton.clicked.connect(lambda : odalarfiyat_paketgüncelle())

guncelle_ui.guncelle_pushButton.clicked.connect(lambda : guncelle_musteriguncelle())

hesap_ui.harcamaEkle_pushButton.clicked.connect(lambda : hesap_harcamaekle(hesap_ui.paketlerTablosu.selectedItems()))
hesap_ui.hesapgetir_pushButton.clicked.connect(lambda : hesap_getir(hesap_ui.musteriNo_lineEdit.text()) )
hesap_ui.ozelharcamaEkle_pushButton.clicked.connect(lambda : hesap_ozelharcamaekle())
hesap_ui.odendi_pushButton.clicked.connect(lambda : hesap_odendi())


rapor_ui.ExceleYaz_pushButton.clicked.connect(lambda : ExceleYaz(rapor_filtre(rapor_ui.giris_dateEdit.date().toPyDate(),
                                                                              rapor_ui.cikis_dateEdit.date().toPyDate(),
                                                                              rapor_ui.kategori_comboBox.currentText(),
                                                                              rapor_ui.blok_comboBox.currentText(),
                                                                              )))

rapor_ui.RaporGetir_pushButton.clicked.connect(lambda : rapor_tabloyaGetir(rapor_filtre(rapor_ui.giris_dateEdit.date().toPyDate(),
                                                                              rapor_ui.cikis_dateEdit.date().toPyDate(),
                                                                              rapor_ui.kategori_comboBox.currentText(),
                                                                              rapor_ui.blok_comboBox.currentText(),
                                                                              )))

sys.exit(Uygulama.exec_())

